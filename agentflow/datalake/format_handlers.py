"""FormatHandler - データフォーマットハンドラ.

各種ファイルフォーマットのパース/シリアライズを担当。

サポートフォーマット:
    - CSV (.csv)
    - JSON (.json, .jsonl)
    - Excel (.xlsx, .xls)
    - Parquet (.parquet)
"""

import csv
import io
import json
import logging
from abc import ABC, abstractmethod
from typing import Any, cast


logger = logging.getLogger(__name__)

# グローバルハンドラレジストリ
_format_handlers: dict[str, "FormatHandler"] = {}


class FormatHandler(ABC):
    """フォーマットハンドラ抽象基底クラス.

    各ファイルフォーマットのパース/シリアライズを実装。
    """

    @property
    @abstractmethod
    def supported_extensions(self) -> set[str]:
        """サポートする拡張子.

        Returns:
            拡張子のセット（例: {'.csv', '.tsv'}）
        """
        ...

    @property
    @abstractmethod
    def supported_content_types(self) -> set[str]:
        """サポートするContent-Type.

        Returns:
            Content-Typeのセット
        """
        ...

    @abstractmethod
    async def parse(self, content: bytes, **kwargs: Any) -> Any:
        """バイナリをパース.

        Args:
            content: バイナリデータ
            **kwargs: パースオプション

        Returns:
            パース結果
        """
        ...

    @abstractmethod
    async def serialize(self, data: Any, **kwargs: Any) -> bytes:
        """データをシリアライズ.

        Args:
            data: シリアライズするデータ
            **kwargs: シリアライズオプション

        Returns:
            バイナリデータ
        """
        ...


class CSVHandler(FormatHandler):
    """CSVハンドラ."""

    @property
    def supported_extensions(self) -> set[str]:
        """サポートする拡張子."""
        return {".csv", ".tsv"}

    @property
    def supported_content_types(self) -> set[str]:
        """サポートするContent-Type."""
        return {"text/csv", "text/tab-separated-values"}

    async def parse(
        self,
        content: bytes,
        encoding: str = "utf-8",
        delimiter: str | None = None,
        **kwargs: Any,
    ) -> list[dict[str, Any]]:
        """CSVをパース.

        Args:
            content: バイナリデータ
            encoding: エンコーディング
            delimiter: 区切り文字（Noneで自動検出）
            **kwargs: csv.DictReaderオプション

        Returns:
            辞書のリスト
        """
        text = content.decode(encoding)
        if delimiter is None:
            # Sniffer で自動検出
            try:
                sample = text[:4096]
                dialect = csv.Sniffer().sniff(sample)
                delimiter = dialect.delimiter
            except csv.Error:
                delimiter = ","

        reader = csv.DictReader(io.StringIO(text), delimiter=delimiter, **kwargs)
        return list(reader)

    async def serialize(
        self,
        data: list[dict[str, Any]],
        encoding: str = "utf-8",
        delimiter: str = ",",
        **kwargs: Any,
    ) -> bytes:
        """CSVにシリアライズ.

        Args:
            data: 辞書のリスト
            encoding: エンコーディング
            delimiter: 区切り文字
            **kwargs: csv.DictWriterオプション

        Returns:
            バイナリデータ
        """
        if not data:
            return b""

        output = io.StringIO()
        fieldnames = list(data[0].keys())
        writer = csv.DictWriter(output, fieldnames=fieldnames, delimiter=delimiter, **kwargs)
        writer.writeheader()
        writer.writerows(data)
        return output.getvalue().encode(encoding)


class JSONHandler(FormatHandler):
    """JSONハンドラ."""

    @property
    def supported_extensions(self) -> set[str]:
        """サポートする拡張子."""
        return {".json", ".jsonl"}

    @property
    def supported_content_types(self) -> set[str]:
        """サポートするContent-Type."""
        return {"application/json", "application/x-ndjson"}

    async def parse(
        self,
        content: bytes,
        encoding: str = "utf-8",
        jsonl: bool = False,
        **kwargs: Any,
    ) -> Any:
        """JSONをパース.

        Args:
            content: バイナリデータ
            encoding: エンコーディング
            jsonl: JSONL形式かどうか
            **kwargs: json.loadsオプション

        Returns:
            パース結果
        """
        text = content.decode(encoding)

        if jsonl or (text.strip().startswith("{") and "\n{" in text):
            # JSONL形式
            return [json.loads(line, **kwargs) for line in text.strip().split("\n") if line.strip()]

        return json.loads(text, **kwargs)

    async def serialize(
        self,
        data: Any,
        encoding: str = "utf-8",
        indent: int | None = 2,
        jsonl: bool = False,
        ensure_ascii: bool = False,
        **kwargs: Any,
    ) -> bytes:
        """JSONにシリアライズ.

        Args:
            data: シリアライズするデータ
            encoding: エンコーディング
            indent: インデント
            jsonl: JSONL形式で出力するか
            ensure_ascii: ASCIIのみ使用するか
            **kwargs: json.dumpsオプション

        Returns:
            バイナリデータ
        """
        if jsonl and isinstance(data, list):
            lines = [json.dumps(item, ensure_ascii=ensure_ascii, **kwargs) for item in data]
            return "\n".join(lines).encode(encoding)

        return json.dumps(data, indent=indent, ensure_ascii=ensure_ascii, **kwargs).encode(encoding)


class ExcelHandler(FormatHandler):
    """Excelハンドラ（openpyxl使用）."""

    @property
    def supported_extensions(self) -> set[str]:
        """サポートする拡張子."""
        return {".xlsx", ".xls"}

    @property
    def supported_content_types(self) -> set[str]:
        """サポートするContent-Type."""
        return {
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "application/vnd.ms-excel",
        }

    async def parse(
        self,
        content: bytes,
        sheet_name: str | int | None = 0,
        **kwargs: Any,
    ) -> list[dict[str, Any]]:
        """Excelをパース.

        Args:
            content: バイナリデータ
            sheet_name: シート名またはインデックス
            **kwargs: 追加オプション

        Returns:
            辞書のリスト
        """
        try:
            import openpyxl
        except ImportError as e:
            msg = "openpyxl is required for Excel support. Install with: pip install openpyxl"
            raise ImportError(msg) from e

        workbook = openpyxl.load_workbook(io.BytesIO(content), read_only=True)

        if isinstance(sheet_name, int):
            sheet = workbook.worksheets[sheet_name]
        elif sheet_name:
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.active

        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []

        headers = [str(h) if h else f"col_{i}" for i, h in enumerate(rows[0])]
        return [dict(zip(headers, row, strict=False)) for row in rows[1:]]

    async def serialize(
        self,
        data: list[dict[str, Any]],
        sheet_name: str = "Sheet1",
        **kwargs: Any,
    ) -> bytes:
        """Excelにシリアライズ.

        Args:
            data: 辞書のリスト
            sheet_name: シート名
            **kwargs: 追加オプション

        Returns:
            バイナリデータ
        """
        try:
            import openpyxl
        except ImportError as e:
            msg = "openpyxl is required for Excel support. Install with: pip install openpyxl"
            raise ImportError(msg) from e

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = sheet_name

        if not data:
            output = io.BytesIO()
            workbook.save(output)
            return output.getvalue()

        # ヘッダー
        headers = list(data[0].keys())
        for col, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col, value=header)

        # データ
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, header in enumerate(headers, 1):
                sheet.cell(row=row_idx, column=col_idx, value=row_data.get(header))

        output = io.BytesIO()
        workbook.save(output)
        return output.getvalue()


class ParquetHandler(FormatHandler):
    """Parquetハンドラ（pyarrow使用）."""

    @property
    def supported_extensions(self) -> set[str]:
        """サポートする拡張子."""
        return {".parquet"}

    @property
    def supported_content_types(self) -> set[str]:
        """サポートするContent-Type."""
        return {"application/vnd.apache.parquet"}

    async def parse(
        self,
        content: bytes,
        columns: list[str] | None = None,
        **kwargs: Any,
    ) -> list[dict[str, Any]]:
        """Parquetをパース.

        Args:
            content: バイナリデータ
            columns: 読み込むカラム
            **kwargs: 追加オプション

        Returns:
            辞書のリスト
        """
        try:
            import pyarrow.parquet as pq
        except ImportError as e:
            msg = "pyarrow is required for Parquet support. Install with: pip install pyarrow"
            raise ImportError(msg) from e

        table = pq.read_table(io.BytesIO(content), columns=columns)
        return cast("list[dict[str, Any]]", table.to_pylist())

    async def serialize(
        self,
        data: list[dict[str, Any]],
        **kwargs: Any,
    ) -> bytes:
        """Parquetにシリアライズ.

        Args:
            data: 辞書のリスト
            **kwargs: 追加オプション

        Returns:
            バイナリデータ
        """
        try:
            import pyarrow as pa
            import pyarrow.parquet as pq
        except ImportError as e:
            msg = "pyarrow is required for Parquet support. Install with: pip install pyarrow"
            raise ImportError(msg) from e

        table = pa.Table.from_pylist(data)
        output = io.BytesIO()
        pq.write_table(table, output)
        return output.getvalue()


def get_format_handler(extension: str) -> FormatHandler | None:
    """拡張子からフォーマットハンドラを取得.

    Args:
        extension: 拡張子（例: '.csv'）

    Returns:
        FormatHandler または None
    """
    ext = extension.lower() if extension.startswith(".") else f".{extension.lower()}"
    return _format_handlers.get(ext)


def register_format_handler(handler: FormatHandler) -> None:
    """フォーマットハンドラを登録.

    Args:
        handler: 登録するハンドラ
    """
    for ext in handler.supported_extensions:
        _format_handlers[ext] = handler
        logger.debug(f"Registered format handler for {ext}")


# デフォルトハンドラを登録
def _register_default_handlers() -> None:
    """デフォルトハンドラを登録."""
    register_format_handler(CSVHandler())
    register_format_handler(JSONHandler())
    register_format_handler(ExcelHandler())
    register_format_handler(ParquetHandler())


_register_default_handlers()
